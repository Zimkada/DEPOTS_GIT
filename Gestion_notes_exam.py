# -*- coding: utf-8 -*-
"""
Created on Wed Jan 24 02:15:10 2024

@author: ok
"""

import tkinter as tk
from tkinter import ttk
from tkinter import messagebox, filedialog, simpledialog
import openpyxl
from openpyxl.styles import NamedStyle

class GestionExamensApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Gestion des Examens")
        self.root.geometry("800x600")
        self.root.configure(bg="#FFD700")  # Couleur orange clair

        self.candidats = []
        self.colleges = set()
        self.centres_composition = set()
        self.moyenne_ponderation = {"Communication": 2, "Lecture": 2, "Histoire-Géographie": 2, "Anglais": 2, "LV2": 2, "PCT": 2, "Mathématiques": 3}

        # Styles pour le tableau
        self.style_header = NamedStyle(name="style_header")
        self.style_header.font.bold = True
        self.style_header.fill = openpyxl.styles.PatternFill(start_color="B0E0E6", end_color="B0E0E6", fill_type="solid")

        # Création des menus
        self.menu_bar = tk.Menu(root)
        self.root.config(menu=self.menu_bar)

        self.menu_importer = tk.Menu(self.menu_bar, tearoff=0)
        self.menu_bar.add_cascade(label="Importer", menu=self.menu_importer)
        self.menu_importer.add_command(label="Importation Excel", command=self.importer_candidats)

        self.menu_consulter = tk.Menu(self.menu_bar, tearoff=0)
        self.menu_bar.add_cascade(label="Consulter", menu=self.menu_consulter)
        self.menu_consulter.add_command(label="Par Collège", command=lambda: self.afficher_liste("College"))
        self.menu_consulter.add_command(label="Par Centre de Composition", command=lambda: self.afficher_liste("CentreComposition"))
        #self.menu_consulter.add_command(label="Liste Globale", command=lambda: self.afficher_liste("Globale"))

        self.menu_attribution_numero = tk.Menu(self.menu_bar, tearoff=0)
        self.menu_bar.add_cascade(label="Attribution Numéro", menu=self.menu_attribution_numero)
        self.menu_attribution_numero.add_command(label="Attribuer Numéro", command=self.attribuer_numero)

        self.menu_numeros_anonymes = tk.Menu(self.menu_bar, tearoff=0)
        self.menu_bar.add_cascade(label="Numéros Anonymes", menu=self.menu_numeros_anonymes)
        self.menu_numeros_anonymes.add_command(label="Générer Numéros Anonymes", command=self.generer_numeros_anonymes)
        self.menu_numeros_anonymes.add_command(label="Afficher Liste", command=self.afficher_liste_numeros_anonymes)

        self.menu_saisir_notes = tk.Menu(self.menu_bar, tearoff=0)
        self.menu_bar.add_cascade(label="Saisir Notes", menu=self.menu_saisir_notes)
        self.menu_saisir_notes.add_command(label="Formulaire de Saisie", command=self.formulaire_saisie_notes)

        self.menu_voir_resultats = tk.Menu(self.menu_bar, tearoff=0)
        self.menu_bar.add_cascade(label="Voir Résultats", menu=self.menu_voir_resultats)
        self.menu_voir_resultats.add_command(label="Afficher Résultats", command=self.afficher_resultats)

        self.menu_imprimer_liste = tk.Menu(self.menu_bar, tearoff=0)
        self.menu_bar.add_cascade(label="Imprimer Liste", menu=self.menu_imprimer_liste)
        self.menu_imprimer_liste.add_command(label="Filtrer par Collège", command=self.imprimer_liste_par_college)

        self.menu_imprimer_bulletin = tk.Menu(self.menu_bar, tearoff=0)
        self.menu_bar.add_cascade(label="Imprimer Bulletin", menu=self.menu_imprimer_bulletin)
        self.menu_imprimer_bulletin.add_command(label="Générer et Imprimer Bulletins", command=self.generer_imprimer_bulletins)

    def importer_candidats(self):
        fichier_excel = filedialog.askopenfilename(filetypes=[("Fichiers Excel", "*.xlsx")])
        if fichier_excel:
            try:
                wb = openpyxl.load_workbook(fichier_excel)
                ws = wb.active

                for row in ws.iter_rows(min_row=2, values_only=True):
                    nom, prenom, sexe, date_naissance, lieu_naissance, option, college, centre_composition = row
                    self.candidats.append({
                        "Nom": nom,
                        "Prenom": prenom,
                        "Sexe": sexe,
                        "DateNaissance": date_naissance,
                        "LieuNaissance": lieu_naissance,
                        "Option": option,
                        "College": college,
                        "CentreComposition": centre_composition,
                        "NumeroInscription": None,
                        "NumeroAnonyme": None,
                        "Notes": {}
                    })

                    self.colleges.add(college)
                    self.centres_composition.add(centre_composition)

                messagebox.showinfo("Importation réussie", "Les candidats ont été importés avec succès.")
            except Exception as e:
                messagebox.showerror("Erreur", f"Erreur lors de l'importation des candidats : {str(e)}")
            finally:
                if wb:
                    wb.close()

    def afficher_liste(self, filtre):
        # Création de la fenêtre Toplevel
        fenetre_liste = tk.Toplevel(self.root)
        fenetre_liste.title("Liste des Candidats")
        fenetre_liste.config(bg="#FFD700")

        # Création d'une liste déroulante pour le filtre
        label_filtre = tk.Label(fenetre_liste, text=f"Sélectionner le {filtre} : ")
        label_filtre.place(x=50, y=20, height=30, width=150)

        options_filtre = list(self.colleges) if filtre == "College" else list(self.centres_composition)
        variable_filtre = tk.StringVar(value=options_filtre[0] if options_filtre else "")
        dropdown_filtre = ttk.Combobox(fenetre_liste, textvariable=variable_filtre, values=options_filtre)
        dropdown_filtre.place(x=250, y=20, height=30, width=200)

        bouton_afficher = tk.Button(fenetre_liste, text="Afficher Liste", command=lambda: self.actualiser_liste(filtre, variable_filtre.get(), fenetre_liste))
        bouton_afficher.place(x=500, y=20, height=30, width=150)
        
        # Création du tableau pour afficher la liste
        colonnes = ["Nom", "Prénoms", "Sexe", "Date de Naissance", "Lieu de Naissance", "Option", "Collège", "Centre de composition"]
        self.tableau = ttk.Treeview(fenetre_liste, columns=colonnes, show="headings", selectmode="none", height=20, )
        self.tableau.place(x=30, y=70, height=600, width=1300)
        #self.tableau.grid(row=1, column=2, columnspan=20, pady=10)
        
        #Création de la barre de défilement verticale
        scrollbar = ttk.Scrollbar(fenetre_liste, orient=tk.VERTICAL, command=self.tableau.yview)
        self.tableau.configure(yscroll=scrollbar.set)
        scrollbar.place(x=5, y=70, height=150, width=20, anchor="nw")
        #scrollbar.grid(row=1, column=1, sticky='ns')

        # Ajout des en-têtes et configuration du style
        for col in colonnes:
            self.tableau.heading(col, text=col)
            self.tableau.column(col, anchor="w", width=160)
            self.tableau.tag_configure("oddrow", background="#FFD700",font=("abadi", 11))  # Fond de couleur orange clair pour les lignes impaires
            self.tableau.tag_configure("evenrow", font=("abadi", 11))

    def actualiser_liste(self, filtre, valeur_filtre, fenetre_liste):
        # Actualise le tableau en fonction du filtre sélectionné
        candidats_filtres = [candidat for candidat in self.candidats if candidat[filtre] == valeur_filtre]

        # Efface toutes les lignes actuelles
        for row in self.tableau.get_children():
            self.tableau.delete(row)

        # Ajoute les nouvelles lignes
        for i, candidat in enumerate(candidats_filtres):
            couleur = "#FFD700" if i % 2 == 0 else ""  # Alternance de couleurs pour une meilleure lisibilité
            self.tableau.insert("", tk.END, values=list(candidat.values()), tags=("oddrow") if couleur else ("evenrow"))
        
       


    def attribuer_numero(self):
        # Attribue les numéros d'inscription aux candidats
        pass

    def generer_numeros_anonymes(self):
        # Génère les numéros anonymes et les attribue aux candidats
        pass

    def afficher_liste_numeros_anonymes(self):
        # Affiche la liste des candidats avec les numéros anonymes
        pass

    def formulaire_saisie_notes(self):
        # Affiche le formulaire de saisie des notes
        pass

    def afficher_resultats(self):
        # Affiche les résultats des candidats
        pass

    def imprimer_liste_par_college(self):
        # Imprime la liste des candidats filtrée par collège
        pass

    def generer_imprimer_bulletins(self):
        # Génère et imprime les bulletins de notes pour chaque candidat
        pass


if __name__ == "__main__":
    root = tk.Tk()
    app = GestionExamensApp(root)
    root.mainloop()
